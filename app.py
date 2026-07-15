import streamlit as st
import pandas as pd
from io import BytesIO
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="IHIP Defaulter & Summary Tool", layout="wide")

# Creating two independent tabs
tab1, tab2 = st.tabs(["Defaulter Analysis", "Reporting Summary"])

# ----------------------------------------------------------------
# TAB 1: DAILY DEFAULTER ANALYSIS (.XLSM OUTPUT)
# ----------------------------------------------------------------
with tab1:
    st.title("Daily IHIP Defaulter Analysis")
    
    # ---------------- UPLOAD SECTION ----------------
    col1, col2, col3 = st.columns(3)
    s_file = col1.file_uploader("S-Form", type=["xlsx"], key="s_def")
    p_file = col2.file_uploader("P-Form", type=["xlsx"], key="p_def")
    l_file = col3.file_uploader("L-Form", type=["xlsx"], key="l_def")
    
    st.markdown("---")
    contact_file = st.file_uploader("Upload Contact File", type=["xlsx"], key="cont_def")
    staff_input = st.text_input("Enter Staff Names (comma separated) e.g. Staff A, Staff B", key="staff_def")
    
    # ---------------- DATE & TIME INPUTS ----------------
    icol1, icol2 = st.columns([2, 3])
    report_date_obj = icol1.date_input("Select Report Date", datetime.date.today(), key="date_input_def")
    formatted_date = report_date_obj.strftime("%d-%m-%Y")
    day_name = report_date_obj.strftime("%A")
    
    icol2.write("Enter Report Time")
    t_c1, t_c2, t_c3 = icol2.columns(3)
    hr_val = t_c1.text_input("HH", value="04", key="hr_t1")
    mn_val = t_c2.text_input("MM", value="05", key="mn_t1")
    am_pm = t_c3.selectbox("AM/PM", ["am", "pm"], index=1, key="ap_t1")
    
    formatted_time = f"{hr_val.zfill(2)}.{mn_val.zfill(2)}{am_pm}"
    report_datetime = f"{day_name} {formatted_date} till {formatted_time}"
    
    # ---------------- OPENPYXL .XLSM GENERATOR ----------------
    def generate_formatted_xlsm(df, title, subtitle):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Defaulter List"
        
        # Define Styles
        font_title = Font(name="Calibri", size=13, bold=True)
        font_sub = Font(name="Calibri", size=11, bold=True)
        font_header = Font(name="Calibri", size=11, bold=True)
        font_cell = Font(name="Calibri", size=11)
        
        fill_title = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        fill_sub = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        fill_header = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        
        thin_side = Side(style='thin', color='A6A6A6')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        align_center = Alignment(horizontal='center', vertical='center')
        
        num_cols = len(df.columns)
        
        # Write Title & Subtitle Row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        ws.cell(row=1, column=1, value=title).font = font_title
        ws.cell(row=1, column=1).fill = fill_title
        ws.cell(row=1, column=1).alignment = align_center
        ws.cell(row=1, column=1).border = border_all
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        ws.cell(row=2, column=1, value=subtitle).font = font_sub
        ws.cell(row=2, column=1).fill = fill_sub
        ws.cell(row=2, column=1).alignment = align_center
        ws.cell(row=2, column=1).border = border_all
        
        # Write Column Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=str(col_name))
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_all
            
        # Write Data Matrix
        for row_idx, row_values in enumerate(df.values, 4):
            for col_idx, val in enumerate(row_values, 1):
                clean_val = "" if pd.isna(val) else val
                cell = ws.cell(row=row_idx, column=col_idx, value=clean_val)
                cell.font = font_cell
                cell.alignment = align_center
                cell.border = border_all
                
        # Set Intentional Grid Layout Column Widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        for col_idx in range(4, num_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
            
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Data Processing Logic
    def process_file(file, form):
        try:
            df = pd.read_excel(file)
            if "facility name" not in str(df.columns).lower():
                for i in range(len(df)):
                    if "facility name" in str(df.iloc[i]).lower():
                        df.columns = df.iloc[i]
                        df = df[i+1:].reset_index(drop=True)
                        break
            df.columns = [str(c).strip() for c in df.columns]
            find_col = lambda k: next((c for c in df.columns if k in str(c).lower()), None)
            
            name = find_col("facility name")
            subtype = find_col("facility sub-type")
            report = find_col("number of times reported")
            ward = next((c for c in df.columns if "ward" in str(c).lower() or "zone" in str(c).lower()), None)
            
            if not (name and subtype and report): return pd.DataFrame()
            
            df[report] = pd.to_numeric(df[report], errors="coerce")
            df = df[df[report].fillna(0) == 0].copy()
            
            category_map = {
                "Dispensary": "PUBLIC", "Government Medical College Hospital": "PUBLIC",
                "IGSL Satellite Laboratory": "PUBLIC", "Infectious Disease Hospital": "PUBLIC",
                "Municipal Hospital": "PUBLIC", "Other Government Hospitals": "PUBLIC",
                "Urban Primary Health Centre": "PUBLIC", "Health Post": "PUBLIC",
                "Health Sub Centre": "PUBLIC", "Private Hospital": "PRIVATE",
                "Private Laboratory": "PRIVATE"
            }
            df["Category"] = df[subtype].map(category_map).fillna("OTHER")
            
            out = pd.DataFrame()
            out["WARD"] = df[ward].fillna("Not Mentioned") if ward else "Not Mentioned"
            out["Facility Name"] = df[name]
            out["Form Type"] = form
            out["Category"] = df["Category"]
            out["REMARK"] = ""
            return out
        except: return pd.DataFrame()

    dfs = []
    if s_file: dfs.append(process_file(s_file, "S FORM"))
    if p_file: dfs.append(process_file(p_file, "P FORM"))
    if l_file: dfs.append(process_file(l_file, "L FORM"))

    if any(not d.empty for d in dfs):
        final_df = pd.concat([d for d in dfs if not d.empty], ignore_index=True)
        final_df["WARD"] = final_df["WARD"].fillna("Not Mentioned").astype(str)
        final_df["ward_sort"] = final_df["WARD"].apply(lambda x: "ZZZ" if x.strip().lower() == "not mentioned" else x)
        final_df = final_df.sort_values(["ward_sort", "Facility Name"]).drop(columns=["ward_sort"])
        
        # Output 1
        out1 = final_df.copy()
        out1.insert(0, "Sr No", range(1, len(out1)+1))
        st.subheader("Output 1")
        st.dataframe(out1, use_container_width=True)
        
        xlsm1 = generate_formatted_xlsm(out1, "IHIP Defaulter", formatted_date)
        st.download_button("Download Output 1 (.xlsm)", xlsm1, f"{formatted_date}_IHIP_Defaulter_List.xlsm", mime="application/vnd.ms-excel.sheet.macroEnabled.12")

        # Output 2
        if staff_input:
            merged = final_df.copy()
            merged["key"] = merged["Facility Name"].astype(str).str.strip().str.lower()
            
            if contact_file:
                cdf = pd.read_excel(contact_file)
                cdf.columns = [str(c).strip() for c in cdf.columns]
                n_c = next((c for c in cdf.columns if "facility" in c.lower()), None)
                p_c = next((c for c in cdf.columns if "contact" in c.lower()), None)
                m_c = next((c for c in cdf.columns if "mobile" in c.lower()), None)
                if n_c and p_c and m_c:
                    cdf["key"] = cdf[n_c].astype(str).str.strip().str.lower()
                    merged = merged.merge(cdf[["key", p_c, m_c]], on="key", how="left")
                    merged.rename(columns={p_c: "Contact Person Name", m_c: "Mobile Number"}, inplace=True)
            
            for col in ["Contact Person Name", "Mobile Number"]:
                if col not in merged.columns: merged[col] = ""
            
            merged["Mobile Number"] = merged["Mobile Number"].astype(str).str.replace(".0", "", regex=False).replace(["nan",""], "Not Available")
            merged["Contact Person Name"] = merged["Contact Person Name"].replace(["nan",""], "Not Available")

            staff_list = [s.strip() for s in staff_input.split(",") if s.strip()]
            n, k = len(merged), len(staff_list)
            if k > 0:
                base, extra = n // k, n % k
                assigned = []
                for i, s in enumerate(staff_list):
                    count = base + (extra if i == k - 1 else 0)
                    assigned.extend([s] * count)
                merged["Assigned Staff"] = assigned
            
            merged.drop(columns=["key"], inplace=True)
            out2 = merged.copy()
            out2.insert(0, "Sr No", range(1, len(out2)+1))
            
            st.markdown("---")
            st.subheader("Output 2")
            st.dataframe(out2, use_container_width=True)
            
            xlsm2 = generate_formatted_xlsm(out2, "IHIP Defaulter List of S, P & L Form", report_datetime)
            st.download_button("Download Output 2 (.xlsm)", xlsm2, f"IHIP_Defaulter_List_{formatted_date}.xlsm", mime="application/vnd.ms-excel.sheet.macroEnabled.12")


# ----------------------------------------------------------------
# TAB 2: CONSOLIDATED REPORTING SUMMARY (.CSV FLAT FORMAT)
# ----------------------------------------------------------------
with tab2:
    st.title("Reporting Summary Status")

    report_date = st.date_input("Select Report Date", datetime.date.today(), key="date_tab2")
    formatted_date = report_date.strftime("%d-%m-%Y")

    sc1, sc2, sc3 = st.columns(3)
    sum_s = sc1.file_uploader("S-Form Summary", type=["xlsx"], key="s_sum")
    sum_p = sc2.file_uploader("P-Form Summary", type=["xlsx"], key="p_sum")
    sum_l = sc3.file_uploader("L-Form Summary", type=["xlsx"], key="l_sum")

    def safe_read_excel(file):
        if file is None: return pd.DataFrame()
        try:
            file.seek(0)
            df = pd.read_excel(file)
            if df.empty: return df
            if all("Unnamed" in str(c) for c in df.columns[:2]):
                df.columns = df.iloc[0]
                df = df[1:].reset_index(drop=True)
            return df
        except Exception as e:
            st.error(f"Read error: {str(e)}")
            return pd.DataFrame()

    def process_summary_file(file, form_name):
        df = safe_read_excel(file)
        if df.empty: return pd.DataFrame()

        raw_cols = {c: str(c).replace(" ", "").lower() for c in df.columns}
        find_fuzzy_col = lambda keywords: next((orig for orig, cln in raw_cols.items() if any(k in cln for k in keywords)), None)

        ward_col = find_fuzzy_col(["admin", "ward"])
        total_col = find_fuzzy_col(["totalreporting"])
        perc_col = find_fuzzy_col(["%ofaverage", "averagereporting"])
        non_rep_col = find_fuzzy_col(["neverreported", "nonreported"])

        if not (ward_col and total_col and perc_col and non_rep_col):
            st.error(f"❌ Missing columns in {form_name}")
            return pd.DataFrame()

        df = df[[ward_col, total_col, perc_col, non_rep_col]].copy()
        df.columns = ["ward", "Total Reporting Units", "% Of Average Reporting Units", "Non Reported Units"]
        df["ward"] = df["ward"].astype(str).str.strip()
        df = df[df["ward"].str.lower() != "nan"]
        
        for col in ["Total Reporting Units", "% Of Average Reporting Units", "Non Reported Units"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    if sum_s and sum_p and sum_l:
        ds = process_summary_file(sum_s, "S-Form")
        dp = process_summary_file(sum_p, "P-Form")
        dl = process_summary_file(sum_l, "L-Form")

        if not ds.empty and not dp.empty and not dl.empty:
            m1 = pd.merge(ds, dp, on="ward", how="outer", suffixes=("_S", "_P"))
            master = pd.merge(m1, dl, on="ward", how="outer")
            
            master.rename(columns={
                "Total Reporting Units": "Total Reporting Units_L",
                "% Of Average Reporting Units": "% Of Average Reporting Units_L",
                "Non Reported Units": "Non Reported Units_L"
            }, inplace=True)

            master = master.fillna(0).sort_values("ward")
            master["Blank1"], master["Blank2"] = "", ""

            final_order = [
                "ward",
                "Total Reporting Units_S", "% Of Average Reporting Units_S", "Non Reported Units_S",
                "Blank1",
                "Total Reporting Units_P", "% Of Average Reporting Units_P", "Non Reported Units_P",
                "Blank2",
                "Total Reporting Units_L", "% Of Average Reporting Units_L", "Non Reported Units_L"
            ]
            
            export_df = master[final_order].copy()
            is_not_mapped = export_df["ward"].str.lower().str.replace(" ", "") == "notmapped"
            main_df = export_df[~is_not_mapped].copy()
            not_mapped_df = export_df[is_not_mapped].copy()

            # Average/Sum Calculation
            sum_data = {"ward": "Total"}
            for col in final_order:
                if col == "ward" or "Blank" in col:
                    sum_data[col] = ""
                    continue
                if "%" in col:
                    sum_data[col] = round(float(main_df[col].mean()), 2)
                elif "Units" in col:
                    sum_data[col] = int(main_df[col].sum())
            
            total_df = pd.DataFrame([sum_data])
            total_df.at[0, "ward"] = "Total"
            
            final_display_df = pd.concat([main_df, total_df, not_mapped_df], ignore_index=True)

            st.subheader("Consolidated Summary Preview")
            st.dataframe(final_display_df, use_container_width=True)

            # --- Flat CSV Transformation ---
            csv_data = final_display_df.to_csv(index=False).encode('utf-8')

            st.download_button(f"📥 Download {formatted_date} Summary Report (.csv)", csv_data, f"{formatted_date}_IHIP_SPL_Status_Report.csv", mime="text/csv")
    else:
        st.info("Upload S, P, and L summary files to begin.")
